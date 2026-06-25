"""Reporting: turn the finished ledger into human-facing outputs.

  * build_sheets(ledger)  -> dict of DataFrames (pure; unit-tested offline)
  * write_excel(...)      -> styled multi-sheet .xlsx (needs openpyxl)
  * write_markdown(...)   -> narrative summary with commentary + alternatives

Tiered presentation (ROADMAP 2): committed assignments are split into
  Assigned    -- unique-in-window or independently corroborated
  Candidates  -- honest ambiguity, shown one row PER CANDIDATE FORMULA
and the unexplained residual keeps its evidence characterization
(iso-partner / has-constraints / isolated) as the below-assignability tier.

The commentary, alternatives, and tier reasons are generated mechanically
from ledger columns, so they are reproducible.
"""
from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd

from . import contexts as X
from . import ledger as L
from . import tiers as T

__version__ = "0.3.0"  # + Below-assignability sheet


def _alts_list(cell) -> list[dict]:
    try:
        v = json.loads(cell) if isinstance(cell, str) else (cell or [])
        return v if isinstance(v, list) else []
    except Exception:
        return []


def _alts_to_text(cell) -> str:
    parts = []
    for a in _alts_list(cell)[:3]:
        s = a.get("ion_score") or a.get("raw_score")
        ppm = a.get("ppm")
        seg = a.get("formula", "?")
        if s is not None:
            seg += f" (score {s:.2f}"
            if ppm is not None:
                seg += f", {ppm:.1f} ppm"
            seg += ")"
        parts.append(seg)
    return "; ".join(parts)


def _iso_to_text(cell) -> str:
    try:
        iso = json.loads(cell) if isinstance(cell, str) else (cell or [])
    except Exception:
        return ""
    return "; ".join(f"{i.get('label')}={i.get('score'):.2f}"
                     for i in iso if i.get("score") is not None)


_RESIDUAL_INTERPRETATION = {
    "iso-partner": ("heavy-isotope satellite of another residual peak -- "
                    "explained the moment its light partner is"),
    "has-constraints": ("isotope structure measured (carbon and/or halogen "
                        "count) -- a constrained formula solve is possible"),
    "isolated": ("no measurable isotope structure -- needs orthogonal "
                 "evidence (e.g. time-series correlation)"),
}


def _enrich_m0(m0: pd.DataFrame) -> pd.DataFrame:
    cls = m0["neutral_formula"].map(lambda f: X.classify_compound(f))
    m0["compound_class"] = [c[0] for c in cls]
    m0["oxidation"] = [c[1] for c in cls]
    m0["heteroatoms"] = [c[2] for c in cls]
    m0["alternatives_text"] = m0["alternatives"].map(_alts_to_text)
    m0["isotopologues_text"] = m0["isotopologues"].map(_iso_to_text)
    return m0


_ASSIGN_COL_ORDER = [
    "mz", "height", "neutral_formula", "adduct", "ion_formula", "dbe",
    "compound_class", "oxidation", "heteroatoms", "ion_score",
    "compound_score", "ppm_error", "confidence", "candidate_density",
    "degeneracy_note", "composite_note", "tier_reason", "isotopologues_text",
    "alternatives_text", "pass_no", "method", "commentary", "peak_id",
]


def _candidate_rows(cand: pd.DataFrame) -> pd.DataFrame:
    """Explode Candidate peaks into one row per candidate formula: rank 1 is
    the committed winner, ranks 2+ are the stored alternatives. This is the
    'stop presenting one formula per peak' sheet."""
    rows = []
    for _, r in cand.iterrows():
        first = {
            "mz": r["mz"], "height": r["height"], "rank": 1,
            "formula": r["neutral_formula"], "adduct": r["adduct"],
            "score": r["ion_score"],
            "eff_score": r.get("eff_score", np.nan),
            "ppm_error": r["ppm_error"], "confidence": r["confidence"],
            "candidate_density": r.get("candidate_density", pd.NA),
            "degeneracy_note": r.get("degeneracy_note", ""),
            "why_candidate": r.get("tier_reason", ""),
            "isotopologues": r.get("isotopologues_text", ""),
            "commentary": r["commentary"], "peak_id": r["peak_id"],
        }
        rows.append(first)
        for k, a in enumerate(_alts_list(r.get("alternatives")), start=2):
            rows.append({
                "mz": r["mz"], "height": r["height"], "rank": k,
                "formula": a.get("formula"), "adduct": a.get("adduct"),
                "score": a.get("raw_score") or a.get("ion_score"),
                "eff_score": a.get("eff_score"),
                "ppm_error": a.get("ppm"),
                "confidence": "", "candidate_density": "",
                "why_candidate": "", "isotopologues": "", "commentary": "",
                "peak_id": r["peak_id"],
            })
    cols = ["mz", "height", "rank", "formula", "adduct", "score", "eff_score",
            "ppm_error", "confidence", "candidate_density", "degeneracy_note",
            "why_candidate", "isotopologues", "commentary", "peak_id"]
    df = pd.DataFrame(rows, columns=cols)
    if len(df):
        df = (df.sort_values(["height", "mz", "rank"],
                             ascending=[False, True, True])
              .reset_index(drop=True))
    return df


def build_sheets(ledger: pd.DataFrame, context: str = "ambient-air",
                 sample_id: str = "") -> dict[str, pd.DataFrame]:
    """Return the report sheets as DataFrames (insertion order == sheet order)."""
    led = ledger.copy()
    # tiering: stamp if the ledger does not carry it (e.g. an old CSV)
    if "tier" not in led.columns or led.loc[led["role"] == L.ROLE_M0, "tier"].isna().all():
        T.apply_tiers(led)
    if "composite_note" not in led.columns:   # old ledgers predate composite detection
        led["composite_note"] = pd.NA
    if "degeneracy_note" not in led.columns:   # old ledgers predate the degeneracy audit
        led["degeneracy_note"] = pd.NA
    m0 = led[led["role"] == L.ROLE_M0].copy()
    if len(m0):
        m0 = _enrich_m0(m0)

    ident = m0[m0["tier"] == T.TIER_ASSIGNED]
    cand = m0[m0["tier"] == T.TIER_CANDIDATE]

    identified = (ident[_ASSIGN_COL_ORDER]
                  .rename(columns={"tier_reason": "evidence"})
                  .sort_values("height", ascending=False)) if len(ident) else \
        pd.DataFrame(columns=[c if c != "tier_reason" else "evidence"
                              for c in _ASSIGN_COL_ORDER])

    candidates = _candidate_rows(cand)

    # unassigned -- characterized by isotope structure (carbon/halogen count,
    # iso-partner class) so the residual is described, not just listed
    from . import residual as RD
    un = RD.characterize_residual(led)
    if len(un):
        un = un.rename(columns={"tier": "evidence"})
        un["interpretation"] = un["evidence"].map(_RESIDUAL_INTERPRETATION)
        un = un[["mz", "height", "evidence", "twin_of", "c_count",
                 "n_Br", "n_Cl", "interpretation", "peak_id"]]

    # by class, with the tier split visible
    if len(m0):
        by_class = (m0.assign(is_ident=(m0["tier"] == T.TIER_ASSIGNED))
                    .groupby(["compound_class", "heteroatoms"])
                    .agg(n_peaks=("peak_id", "count"),
                         n_identified=("is_ident", "sum"),
                         signal=("height", "sum"))
                    .reset_index())
        by_class["n_candidates"] = by_class["n_peaks"] - by_class["n_identified"]
        by_class = (by_class[["compound_class", "heteroatoms", "n_peaks",
                              "n_identified", "n_candidates", "signal"]]
                    .sort_values("signal", ascending=False))
    else:
        by_class = pd.DataFrame()

    # unique formulas across channels
    if len(m0):
        uniq = (m0.assign(is_ident=(m0["tier"] == T.TIER_ASSIGNED))
                .groupby("neutral_formula")
                .agg(n_peaks=("peak_id", "count"),
                     adducts=("adduct", lambda s: "; ".join(sorted(set(map(str, s))))),
                     best_tier=("is_ident", lambda s: T.TIER_ASSIGNED if s.any()
                                else T.TIER_CANDIDATE),
                     best_score=("ion_score", "max"),
                     signal=("height", "sum"))
                .reset_index().sort_values("signal", ascending=False))
    else:
        uniq = pd.DataFrame()

    # isotopologues, joined back to the parent so each row is self-describing
    iso = led[led["role"] == L.ROLE_ISO][[
        "peak_id", "mz", "height", "parent_peak_id", "iso_label",
        "iso_match_score"]].copy()
    if len(iso) and len(m0):
        parents = m0.set_index("peak_id")
        iso["parent_formula"] = iso["parent_peak_id"].map(parents["neutral_formula"])
        iso["parent_adduct"] = iso["parent_peak_id"].map(parents["adduct"])
        iso["parent_mz"] = iso["parent_peak_id"].map(parents["mz"])
        iso = (iso[["mz", "height", "iso_label", "iso_match_score",
                    "parent_formula", "parent_adduct", "parent_mz",
                    "parent_peak_id", "peak_id"]]
               .sort_values(["parent_mz", "mz"]))

    # ownership audit (one row per physical peak)
    ownership = led[["peak_id", "mz", "height", "role", "tier",
                     "neutral_formula", "adduct", "ion_score", "ppm_error",
                     "confidence", "composite_note", "parent_peak_id",
                     "iso_label", "pass_no", "method", "commentary"]
                    ].sort_values("height", ascending=False)

    # target list (formula + adduct + best ppm), Assigned first
    # ('Assigned' < 'Candidate' lexically, hence the ascending tier sort)
    target = (m0[["neutral_formula", "adduct", "ion_formula", "mz",
                  "ppm_error", "ion_score", "confidence", "tier"]]
              .sort_values(["tier", "mz"], ascending=[True, True])) if len(m0) else pd.DataFrame()

    reag = led[led["role"] == L.ROLE_REAGENT][[
        "mz", "height", "commentary", "peak_id"]].copy().sort_values(
        "height", ascending=False)

    # below-assignability: M0 commits flagged as mass-saturated O-monsters -- the
    # base mass fits but ~dozens of plausible ions sit within <=1 ppm, so the
    # formula is one arbitrary pick, NOT an identification. Listed as a constrained
    # mass + the tie-set size, separated from the real Candidates.
    if "below_assignability" in led.columns:
        bamask = (led["role"] == L.ROLE_M0) & led["below_assignability"].fillna(False).astype(bool)
        bcols = [c for c in ["mz", "neutral_formula", "adduct", "ion_formula", "ppm_error",
                             "ion_score", "degeneracy_density", "degeneracy_note", "tier_reason"]
                 if c in led.columns]
        below = led[bamask][bcols].copy().sort_values("mz") if bamask.any() else pd.DataFrame(columns=bcols)
    else:
        below = pd.DataFrame()

    return {
        "Summary": summary_stats(led, context=context, sample_id=sample_id),
        "Read me": legend_sheet(),
        "Assigned": identified,
        "Candidates": candidates,
        "Below assignability": below,
        "Unassigned": un,
        "By class": by_class,
        "Unique formulas": uniq,
        "Isotopologues": iso,
        "Peak ownership": ownership,
        "Target list": target,
        "Reagent ions": reag,
    }


def summary_stats(ledger: pd.DataFrame, *, context: str = "",
                  sample_id: str = "") -> pd.DataFrame:
    st = L.stats(ledger)
    n = st["n_peaks"]
    rows = []

    def add(section, metric, value):
        rows.append({"section": section, "metric": metric, "value": value})

    if sample_id:
        add("Run", "sample_id", sample_id)
    if context:
        add("Run", "context", context)
    # NB: no "generated (UTC)" cell here on purpose — this workbook is MATERIAL DATA
    # and must be byte-identical for identical inputs regardless of when it's run.
    # The run timestamp lives only on the PDF report cover + the run-folder name.
    add("Run", "peaks total", n)

    role_label = {L.ROLE_M0: "M0 (has formula)", L.ROLE_ISO: "isotopologue children",
                  L.ROLE_REAGENT: "reagent ions",
                  L.ROLE_UNEXPLAINED: "unexplained"}
    for role, label in role_label.items():
        cnt = st["by_role"].get(role, 0)
        add("Coverage", label,
            f"{cnt}  ({100 * st['count_frac_by_role'].get(role, 0):.1f}% of peaks, "
            f"{100 * st['signal_by_role'].get(role, 0):.1f}% of signal)")
    expl = (st["signal_by_role"].get(L.ROLE_M0, 0)
            + st["signal_by_role"].get(L.ROLE_ISO, 0)
            + st["signal_by_role"].get(L.ROLE_REAGENT, 0))
    add("Coverage", "signal explained", f"{100 * expl:.1f}%")

    m0 = ledger[ledger["role"] == L.ROLE_M0]
    if "tier" in ledger.columns and len(m0):
        h_m0 = m0["height"].sum(skipna=True)
        for tier in (T.TIER_ASSIGNED, T.TIER_CANDIDATE):
            sub = m0[m0["tier"] == tier]
            sig = (100 * sub["height"].sum(skipna=True) / h_m0) if h_m0 else 0.0
            add("Tiers", tier,
                f"{len(sub)}  ({100 * len(sub) / len(m0):.0f}% of assignments, "
                f"{sig:.0f}% of assigned signal)")
        add("Tiers", "Below assignability",
            f"{st['by_role'].get(L.ROLE_UNEXPLAINED, 0)} unexplained peaks "
            "(see Unassigned sheet for per-peak evidence)")

    if len(m0):
        base = m0["confidence"].map(T.base_confidence)
        for lab in ("High", "Good", "Low", "Suspect"):
            cnt = int((base == lab).sum())
            if cnt:
                add("Confidence", lab, cnt)
        for meth, cnt in m0["method"].value_counts().items():
            add("Methods", str(meth), int(cnt))
    return pd.DataFrame(rows, columns=["section", "metric", "value"])


def legend_sheet() -> pd.DataFrame:
    rows = [
        ("Tiers", "Assigned", "Formula unique in the calibrated mass window, "
         "or corroborated by independent evidence: Mascope-confirmed "
         "isotopologues, the same neutral in a second ionization channel, or "
         "series-anchor support. The 'evidence' column states which."),
        ("Tiers", "Candidate", "A plausible formula that the evidence cannot "
         "single out. All competing formulas are listed, one row per candidate "
         "(rank 1 = the committed best guess). The 'why_candidate' column "
         "gives the demotion reason."),
        ("Tiers", "Below assignability", "Unexplained peaks (Unassigned sheet)."
         " 'evidence' says what the isotope pattern DOES tell us: iso-partner "
         "(satellite of another residual peak), has-constraints (carbon/"
         "halogen count measured), isolated (no isotope structure)."),
        ("Confidence", "High", "Score >= tau_high, |ppm| within 1.5x the gate, "
         "at least one Mascope-confirmed isotopologue, no near-tie."),
        ("Confidence", "Good", "Score >= tau_good and |ppm| within 2x the gate. "
         "Pattern-driven passes (series, iso-pair) are capped here by design."),
        ("Confidence", "Low / Suspect", "Mass fit only, or pattern evidence "
         "with a weaker score. Always tier Candidate."),
        ("Columns", "height", "Peak intensity (cps) from the summed spectrum."),
        ("Columns", "ion_score / compound_score", "Mascope match_compounds "
         "scores for the matched ion / whole compound (isotope pattern "
         "included). Server scoring is authoritative."),
        ("Columns", "eff_score", "Arbitration score: raw score minus the "
         "complexity prior (heteroatom skepticism, waived when the diagnostic "
         "isotope is confirmed) and minor-channel penalty."),
        ("Columns", "ppm_error", "Observed minus theoretical m/z, ppm. The "
         "pipeline self-calibrates (mu, sigma) on the pass-1 CHO/CHON "
         "backbone and gates later commits by z-score."),
        ("Columns", "candidate_density", "How many formulas live within 0.10 "
         "effective score of the winner (winner included). 1 = unique. "
         "'>=N' means the stored alternatives list saturated."),
        ("Columns", "dbe", "Double-bond equivalents of the NEUTRAL (can be "
         "half-integer for radicals, e.g. HO2)."),
        ("Columns", "isotopologues", "Mascope-scored isotope satellites "
         "attributed to this assignment (label=score)."),
        ("Methods", "known:*", "Pass 0: locked list of known instrument "
         "contaminants (silanediol/PDMS ladder) and small atmospheric "
         "acids/radicals; mass + own-81Br-twin self-consistency gated."),
        ("Methods", "cheminfo+grid", "Pass 1: CHO/CHON backbone candidates "
         "from the cheminfo m/z query + local formula grid, scored by "
         "Mascope; self-calibration is fitted on these."),
        ("Methods", "gka-series", "Pass 2: generalized-Kendrick homologous "
         "series extension of assigned anchors."),
        ("Methods", "contaminant:* / cluster:HBr", "Pass 3: evidence-opened "
         "contaminant families (repeat units validated against decoys) and "
         "HBr cluster reading of reagent-halogen compositions."),
        ("Methods", "residual:iso-pair", "Pass 4: Br/Cl/BrCl isotope doublets "
         "in the residual anchor a constrained enumeration (carbon-clamped "
         "by the 13C satellite where measurable)."),
        ("Methods", "residual:series", "Pass 4: bright residual peaks 1-2 "
         "exact repeat units from >= 1 assigned anchors."),
        ("Methods", "completion:known-neutral", "Pass 5: cross-channel "
         "partners + series-gap fills of already-assigned neutrals "
         "(no new formula space)."),
        ("Provenance", "ledger.csv", "The per-peak ledger is the source of "
         "truth; every sheet here is a mechanical view of it. Commentary "
         "strings are generated from ledger columns and are reproducible."),
    ]
    return pd.DataFrame(rows, columns=["section", "topic", "explanation"])


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------
_NUM_FMT = {
    "mz": "0.0000", "parent_mz": "0.0000",
    "height": "#,##0", "signal": "#,##0",
    "ppm_error": "+0.00;-0.00;0.00",
    "ion_score": "0.000", "compound_score": "0.000", "eff_score": "0.000",
    "score": "0.000", "best_score": "0.000", "iso_match_score": "0.000",
    "dbe": "0.0",
}
_WRAP_COLS = {"commentary": 70, "evidence": 46, "why_candidate": 46,
              "tier_reason": 46, "alternatives_text": 44, "composite_note": 50,
              "degeneracy_note": 60,
              "isotopologues_text": 30, "isotopologues": 30,
              "interpretation": 52, "explanation": 90, "value": 46}

_FILL = {
    "good":    ("C6EFCE", "006100"),
    "okay":    ("E2EFDA", "375623"),
    "warn":    ("FFEB9C", "9C6500"),
    "bad":     ("FFC7CE", "9C0006"),
    "info":    ("DDEBF7", "1F4E79"),
    "neutral": ("EDEDED", "3B3838"),
}


def _chip(label: str) -> str | None:
    s = str(label)
    if s == T.TIER_ASSIGNED or s.startswith("High"):
        return "good"
    if s == T.TIER_CANDIDATE or s.startswith("Low"):
        return "warn"
    if s.startswith("Good"):
        return "okay"
    if s.startswith("Suspect"):
        return "bad"
    if s == "iso-partner":
        return "info"
    if s == "has-constraints":
        return "warn"
    if s == "isolated":
        return "neutral"
    return None


def _style_sheet(ws, df, *, chip_cols=(), band_by=None):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for j in range(1, len(df.columns) + 1):
        c = ws.cell(row=1, column=j)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    if len(df):
        ws.auto_filter.ref = ws.dimensions

    # column widths + number formats + wrapping
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        if col in _WRAP_COLS:
            width = _WRAP_COLS[col]
        else:
            lens = df[col].head(200).map(lambda v: len(str(v)))
            width = max(int(lens.max()) if len(lens) else 8, len(str(col)) + 1)
            width = min(width + 2, 26)
        ws.column_dimensions[letter].width = width
        fmt = _NUM_FMT.get(col)
        wrap = col in _WRAP_COLS
        if fmt or wrap:
            align = Alignment(wrap_text=True, vertical="top") if wrap else None
            for i in range(2, len(df) + 2):
                cell = ws.cell(row=i, column=j)
                if fmt:
                    cell.number_format = fmt
                if align:
                    cell.alignment = align

    # colour chips on verdict-like columns
    for col in chip_cols:
        if col not in df.columns:
            continue
        j = list(df.columns).index(col) + 1
        for i, val in enumerate(df[col].tolist(), start=2):
            kind = _chip(val) if pd.notna(val) else None
            if kind:
                bg, fg = _FILL[kind]
                cell = ws.cell(row=i, column=j)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(color=fg)

    # alternating banding by a grouping key (Candidates: one band per peak)
    if band_by is not None and band_by in df.columns and len(df):
        band_fill = PatternFill("solid", fgColor="F2F2F2")
        bold = Font(bold=True)
        prev, band = None, False
        rank_j = (list(df.columns).index("rank") + 1) if "rank" in df.columns else None
        formula_j = (list(df.columns).index("formula") + 1) if "formula" in df.columns else None
        for i, key in enumerate(df[band_by].tolist(), start=2):
            if key != prev:
                band = not band
                prev = key
            if band:
                for j in range(1, len(df.columns) + 1):
                    if ws.cell(row=i, column=j).fill.start_color.rgb in (None, "00000000"):
                        ws.cell(row=i, column=j).fill = band_fill
            if rank_j and formula_j and ws.cell(row=i, column=rank_j).value == 1:
                ws.cell(row=i, column=formula_j).font = bold


def _style_summary(ws, df):
    """Section-grouped key/value look for Summary and Read me."""
    from openpyxl.styles import Border, Font, Side
    top = Border(top=Side(style="thin", color="B0B0B0"))
    bold = Font(bold=True)
    prev = None
    sec_j = 1
    for i, val in enumerate(df.iloc[:, 0].tolist(), start=2):
        if val != prev:
            for j in range(1, len(df.columns) + 1):
                ws.cell(row=i, column=j).border = top
            ws.cell(row=i, column=sec_j).font = bold
            prev = val
        else:
            ws.cell(row=i, column=sec_j).value = None


def write_excel(ledger: pd.DataFrame, path: str | Path,
                context: str = "ambient-air", sample_id: str = ""):
    sheets = build_sheets(ledger, context, sample_id)
    chip_cols = ("tier", "confidence", "evidence", "best_tier")
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        for name, df in sheets.items():
            out = df if len(df) else pd.DataFrame({"(empty)": []})
            out.to_excel(xl, sheet_name=name[:31], index=False)
            ws = xl.sheets[name[:31]]
            if not len(df):
                continue
            _style_sheet(ws, out,
                         chip_cols=[c for c in chip_cols if c in out.columns],
                         band_by="peak_id" if name == "Candidates" else None)
            if name in ("Summary", "Read me"):
                _style_summary(ws, out)
    # content-stable bytes (fixed SOURCE_DATE_EPOCH) so the assignment workbook is a
    # pure function of the ledger, matching the cluster workbook.
    from .cluster import _make_xlsx_deterministic
    _make_xlsx_deterministic(path)
    return path


def write_markdown(result: dict, path: str | Path) -> Path:
    led = result["ledger"]
    st = result["stats"]
    m0 = led[led["role"] == L.ROLE_M0]
    tier_line = ""
    if "by_tier" in st and st["by_tier"]:
        tier_line = "  |  ".join(f"{k}: {v}" for k, v in st["by_tier"].items())
    lines = [
        f"# Peak assignment — sample {result['sample_id']}",
        "",
        f"- Context: **{result['context']}**",
        f"- Peaks: {st['n_peaks']}  |  M0 (has formula): {st['by_role']['M0']}  "
        f"|  isotopologues: {st['by_role']['iso_child']}  "
        f"|  unexplained: {st['by_role']['unexplained']}",
        (f"- Peaks explained: "
         f"{100*(1 - st['count_frac_by_role']['unexplained']):.1f}% "
         f"({st['by_role']['unexplained']}/{st['n_peaks']} unexplained)  |  "
         if "count_frac_by_role" in st else "- ")
        + f"Signal explained: "
        f"{100*(st['signal_by_role']['M0']+st['signal_by_role']['iso_child']+st['signal_by_role']['reagent']):.1f}%",
    ]
    if tier_line:
        lines.append(f"- Tiers: {tier_line}")
    lines += [
        f"- Confidence: {st.get('by_confidence', {})}",
        f"- Prescan: {result['prescan']}",
        "",
        "## Top assignments",
        "",
    ]
    top = m0.sort_values("ion_score", ascending=False).head(20)
    for _, r in top.iterrows():
        tier = f" [{r['tier']}]" if "tier" in m0.columns and pd.notna(r.get("tier")) else ""
        lines.append(f"- **{r['neutral_formula']}** {r['adduct']} "
                     f"(m/z {r['mz']:.4f}, {r['confidence']}{tier}) — {r['commentary']}")
    if result.get("problems"):
        lines += ["", "## ⚠ Ledger validation problems", ""]
        lines += [f"- {p}" for p in result["problems"]]
    Path(path).write_text("\n".join(lines))
    return Path(path)
